# 📁 backend/python_scripts/extraccion_texto.py
import mysql.connector
import pandas as pd
import io, os, re
from dotenv import load_dotenv

from PyPDF2 import PdfReader
import docx
import textract
import mammoth
import openpyxl

load_dotenv()

connection = mysql.connector.connect(
    host=os.getenv("DB_HOST"),
    user=os.getenv("DB_USER"),
    password=os.getenv("DB_PASSWORD"),
    database=os.getenv("DB_NAME")
)
cursor = connection.cursor(dictionary=True)


def limpiar_texto(texto):
    texto = re.sub(r"\s+", " ", texto)  # Quitar espacios dobles
    texto = re.sub(r"[^\w\sáéíóúÁÉÍÓÚñÑ.,;:!?%()-]", "", texto)
    return texto.strip()


def extraer_texto(tipo, blob):
    try:
        data = io.BytesIO(blob)
        texto = ""

        if tipo in ["PDF"]:
            reader = PdfReader(data)
            for page in reader.pages:
                texto += page.extract_text() or ""

        elif tipo in ["DOCX"]:
            doc = docx.Document(data)
            texto = "\n".join([p.text for p in doc.paragraphs])

        elif tipo in ["TXT", "RTF", "HTML", "ODT"]:
            texto = data.read().decode("utf-8", errors="ignore")

        elif tipo in ["XLSX"]:
            wb = openpyxl.load_workbook(data, read_only=True)
            for sheet in wb.sheetnames:
                hoja = wb[sheet]
                for row in hoja.iter_rows(values_only=True):
                    texto += " ".join([str(c) for c in row if c]) + " "

        else:
            return None  # Ignorar tipos no textuales

        return limpiar_texto(texto)
    except Exception as e:
        print(f"Error al procesar {tipo}: {e}")
        return None


cursor.execute("SELECT id_recurso, tipo, archivo FROM recurso WHERE origen_dato='NUEVO'")
recursos = cursor.fetchall()

procesados = 0
for r in recursos:
    texto = extraer_texto(r["tipo"].upper(), r["archivo"])
    if texto and len(texto) > 30:
        cursor.execute("""
            INSERT INTO recurso_texto (id_recurso, texto_limpio)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE texto_limpio = VALUES(texto_limpio);
        """, (r["id_recurso"], texto))
        procesados += 1
        print(f"✅ Texto extraído de recurso ID {r['id_recurso']} ({r['tipo']})")

connection.commit()
print(f"\n🔹 Total recursos procesados: {procesados}")

cursor.close()
connection.close()
